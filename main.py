"""
main.py
=======
AISaleAnalyst — main entry point.

Run
---
    python main.py

The script prompts for an estate-sale listing URL, downloads images via
:mod:`scrapers.ListingExtractor`, analyses each image with the configured
AI model, deduplicates results, fetches eBay sold-listing comps, and
writes a self-contained HTML report.

Pipeline
--------
1. **Vision pass** — :func:`core.vision.analyze_image` sends each image
   to the AI and returns a structured item dict.
2. **Deduplication** — :func:`core.deduplication.deduplicate` runs a
   fuzzy pass then an optional AI-powered pass to collapse duplicate
   photos of the same physical object.
3. **eBay comps** — :func:`core.ebay.scrape_ebay_comps` fetches
   sold-listing prices per unique item using a 3-level progressive
   fallback.
4. **Report** — :func:`core.report.generate_report` writes the HTML
   report ranked by ROI.

Project layout
--------------
::

    AISaleAnalyst/
    ├── main.py             ← you are here
    ├── core/               ← business logic
    │   ├── config.py
    │   ├── vision.py
    │   ├── deduplication.py
    │   ├── ebay.py
    │   ├── financials.py
    │   └── report.py
    └── scrapers/           ← site-specific downloaders
        ├── ListingExtractor.py
        ├── EstateSalesNet.py
        ├── EstateSalesOrg.py
        └── MaxSold.py
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import time
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

from core.config import IMAGES_FOLDER, MAX_IMAGES, VISION_WORKERS, EBAY_WORKERS, OUTPUT_FOLDER, image_to_base64, USE_DEDUP, GENERATE_DUPLICATES_REPORT, DEV_MODE
from core.deduplication import deduplicate, post_dedup_verify
from core.ebay import scrape_ebay_comps, close_ebay_session
from core.report import generate_report
from core.vision import analyze_image
from scrapers.ListingExtractor import identifySite

# ---------------------------------------------------------------------------
# Duplicates Excel report generator
# ---------------------------------------------------------------------------


def _generate_duplicates_xlsx(
    items: list,
    merge_log: list[dict],
    similar_flags: list[dict],
    output_path: str,
) -> None:
    """
    Generate a Duplicates & Probable Duplicates Excel report.

    Creates four worksheets:
    - **Exact Duplicates**: Items with identical names (post-dedup).
    - **Probable Duplicates**: Items with similar but not identical names.
    - **AI Verified Merges**: Groups that were merged after visual verification.
    - **Flagged Similar**: Items flagged as similar but kept separate.
    """
    from collections import defaultdict
    import re

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("  [duplicates] openpyxl not installed — skipping Excel report.")
        print("  [duplicates] Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    def _style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # --- Sheet 1: Exact Duplicates (items with identical names) ---
    ws1 = wb.active
    ws1.title = "Exact Duplicates"
    ws1.append(["Item", "Occurrences"])

    name_counts: dict[str, int] = defaultdict(int)
    for item in items:
        name = item["ai"].get("item_name", "Unknown")
        name_counts[name] += 1

    for name, count in sorted(name_counts.items(), key=lambda x: x[1], reverse=True):
        if count >= 2:
            ws1.append([name, count])

    _style_header(ws1)
    ws1.column_dimensions["A"].width = 50
    ws1.column_dimensions["B"].width = 14

    # --- Sheet 2: Probable Duplicates (fuzzy name matches) ---
    ws2 = wb.create_sheet("Probable Duplicates")
    ws2.append(["Normalized Group", "Representative Names", "Total Occurrences"])

    from core.deduplication import _deep_normalize, _similarity

    norm_groups: dict[str, list[str]] = defaultdict(list)
    for item in items:
        name = item["ai"].get("item_name", "Unknown")
        norm = _deep_normalize(name)
        if norm:
            norm_groups[norm].append(name)

    # Find normalized groups that contain items with different original names
    for norm_key, names in sorted(norm_groups.items(), key=lambda x: len(x[1]), reverse=True):
        unique_names = list(set(names))
        if len(unique_names) >= 2:
            ws2.append([norm_key, ", ".join(unique_names), len(names)])

    _style_header(ws2)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 70
    ws2.column_dimensions["C"].width = 18

    # --- Sheet 3: AI Verified Merges ---
    ws3 = wb.create_sheet("AI Verified Merges")
    ws3.append(["Merged Items", "Items Merged", "Surviving Item"])

    for entry in merge_log:
        names = entry.get("names", [])
        surviving_idx = entry.get("survivor_idx", "?")
        ws3.append([
            ", ".join(names),
            entry.get("merged_count", len(names)),
            names[0] if names else "?",
        ])

    _style_header(ws3)
    ws3.column_dimensions["A"].width = 70
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 40

    # --- Sheet 4: Flagged Similar ---
    ws4 = wb.create_sheet("Flagged Similar")
    ws4.append(["Similar Items", "Count", "Status"])

    for entry in similar_flags:
        names = entry.get("names", [])
        status = "Kept separate"
        if entry.get("error"):
            status = f"Error: {entry['error'][:50]}"
        elif entry.get("mode") == "flag_only":
            status = "Flag only (visual verify disabled)"
        ws4.append([", ".join(names), len(names), status])

    _style_header(ws4)
    ws4.column_dimensions["A"].width = 70
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 40

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------


def main(max_images_override: int | None = None) -> None:
    """
    Run the full AISaleAnalyst pipeline end-to-end.

    Steps
    -----
    1. Resolve the images folder (prompt user if not set in :mod:`core.config`).
    2. Collect image file paths up to ``MAX_IMAGES``.
    3. AI vision pass — analyse each image.
    4. Deduplication — fuzzy + optional AI pass.
    5. eBay comps pass — one search per unique item.
    6. HTML report generation.
    """
    from pathlib import Path

    effective_max = max_images_override if max_images_override is not None else MAX_IMAGES
    if DEV_MODE:
        effective_max = min(effective_max, 20)

    if DEV_MODE:
        print("\n" + "="*60)
        print("⚠️  DEV_MODE IS ACTIVE  ⚠️")
        print("="*60)
        print(f"• Items capped to: {effective_max} (DEV_MODE max is 20)")
        print(f"• Emails marked as DEV_MODE test")
        print("• Model swapped to cheaper alternative")
        print("="*60 + "\n")

    # --- Resolve images folder
    images_folder = IMAGES_FOLDER
    url = None
    if images_folder is None:
        url           = input("Enter Estate listing URL: ").strip()
        
        # Determine the target output folder based on URL domain
        if "estatesales.net" in url:
            target_folder = "EstateSaleNetOutput"
        elif "estatesales.org" in url:
            target_folder = "EstateSalesOrgOutput"
        elif "maxsold.com" in url:
            target_folder = "MaxSoldOutput"
        else:
            raise ValueError(f"Unsupported URL domain: '{url}'. Supported platforms are EstateSales.net, EstateSales.org, and MaxSold.com")
            
        import shutil
        if Path(target_folder).exists():
            print(f"Clearing old data from '{target_folder}'...")
            shutil.rmtree(target_folder, ignore_errors=True)
            
        images_folder = identifySite(url, max_images=effective_max)
        
        last_url_file = Path(target_folder) / "last_url.txt"
        # Save the last URL to the directory for future runs
        try:
            Path(images_folder).mkdir(parents=True, exist_ok=True)
            last_url_file.write_text(url, encoding="utf-8")
        except Exception as e:
            print(f"Warning: Could not save last URL metadata: {e}")
    else:
        last_url_file = Path(images_folder) / "last_url.txt"
        if last_url_file.exists():
            try:
                url = last_url_file.read_text(encoding="utf-8").strip()
            except Exception:
                pass

    folder      = Path(images_folder)
    extensions  = {".jpg", ".jpeg", ".png", ".webp"}
    
    image_files = sorted(
        f for f in folder.iterdir() if f.suffix.lower() in extensions
    )[:effective_max]

    if not image_files:
        print(f"No images found in {images_folder}")
        return

    print(f"Found {len(image_files)} images. Starting analysis...\n")

    # --- Check for existing progress file
    progress_file = folder / "vision_progress.json"
    cached_data = {}
    if progress_file.exists():
        ans = input(f"Found existing progress file '{progress_file.name}'. Resume from previous run? (y/n) [n]: ").strip().lower()
        if ans == 'y':
            try:
                with open(progress_file, "r", encoding="utf-8") as f:
                    progress_list = json.load(f)
                    for item in progress_list:
                        # Use resolved path string as cache key
                        cached_data[str(Path(item["image"]).resolve())] = item
                print(f"Resuming run. Loaded {len(cached_data)} cached image analyses.")
            except Exception as e:
                print(f"Error reading progress file, starting fresh: {e}")

    raw_results: list[dict] = []
    skipped_results: list[dict] = []
    
    # Process cached entries
    images_to_analyze = []
    for img_path in image_files:
        resolved_path_str = str(img_path.resolve())
        if resolved_path_str in cached_data:
            item = cached_data[resolved_path_str]
            if item.get("ai", {}).get("skip"):
                skipped_results.append(item)
            else:
                raw_results.append(item)
        else:
            images_to_analyze.append(img_path)

    # If there are new images to analyze, process them in parallel
    if images_to_analyze:
        print(f"Starting parallel analysis of {len(images_to_analyze)} remaining images with {VISION_WORKERS} workers...\n")
        
        results_lock = threading.Lock()
        counter = len(cached_data)
        total_images = len(image_files)

        def process_image(img_path, index):
            nonlocal counter
            # Stagger startup times slightly to avoid immediate rate limit spikes
            time.sleep((index % VISION_WORKERS) * 0.15)
            
            ai_result = analyze_image(str(img_path))
            thumb_data = image_to_base64(str(img_path))
            
            with results_lock:
                counter += 1
                if ai_result.get("skip"):
                    skip_reason = ai_result.get("skip_reason", "Unknown reason")
                    print(f"[{counter}/{total_images}] {img_path.name} -> Skipped ({skip_reason})")
                    skipped_results.append({
                        "image": str(img_path),
                        "ai":    ai_result,
                        "thumb": thumb_data,
                    })
                else:
                    pkg_l = ai_result.get("pkg_length_in", 0)
                    pkg_w = ai_result.get("pkg_width_in", 0)
                    pkg_h = ai_result.get("pkg_height_in", 0)
                    pkg_wt = ai_result.get("pkg_weight_lb", 0)
                    print(
                        f"[{counter}/{total_images}] {img_path.name} -> "
                        f"{ai_result.get('item_name')} "
                        f"| group: {ai_result.get('item_group')} "
                        f"| {ai_result.get('confidence')}% "
                        f"| pkg: {pkg_l}x{pkg_w}x{pkg_h} in, {pkg_wt} lbs"
                    )
                    raw_results.append({
                        "image": str(img_path),
                        "ai":    ai_result,
                        "thumb": thumb_data,
                    })
                
                # Save progress incrementally to disk
                all_progress = raw_results + skipped_results
                try:
                    with open(progress_file, "w", encoding="utf-8") as f:
                        json.dump(all_progress, f, indent=2)
                except Exception as e:
                    print(f"  [Warning] Failed to write progress file: {e}")

        import queue
        task_queue = queue.Queue()
        for idx, img_path in enumerate(images_to_analyze):
            task_queue.put((idx, img_path))
            
        def worker():
            while True:
                try:
                    idx, img_path = task_queue.get_nowait()
                except queue.Empty:
                    break
                try:
                    process_image(img_path, idx)
                except Exception as exc:
                    print(f"  Thread exception: {exc}")
                finally:
                    task_queue.task_done()
                    
        threads = []
        for _ in range(VISION_WORKERS):
            t = threading.Thread(target=worker)
            t.daemon = True
            t.start()
            threads.append(t)
            
        try:
            for t in threads:
                while t.is_alive():
                    t.join(0.5)
        except KeyboardInterrupt:
            print("\n[!] Ctrl+C detected! Shutting down vision workers...")
            # Empty the queue so threads stop
            while not task_queue.empty():
                try:
                    task_queue.get_nowait()
                except queue.Empty:
                    break
            import sys
            sys.exit(1)
    else:
        print("All images loaded from cache. No new analysis needed.")

    if not raw_results:
        print("No items identified from images.")
        return

    # --- Step 2: Deduplication
    if USE_DEDUP:
        print("Running deduplication...")
        unique_results = deduplicate(raw_results)
    else:
        print("Deduplication bypassed (USE_DEDUP=False).")
        unique_results = raw_results

    # --- Step 2b: Post-dedup name-similarity + visual verification
    unique_results, merge_log, similar_flags = post_dedup_verify(unique_results)

    # --- Step 3: eBay comps (Multi-threaded HTTP workers)
    print(f"\nFetching eBay comps for {len(unique_results)} unique items across {EBAY_WORKERS} parallel workers...\n")
    
    comp_counter = 0
    total_unique = len(unique_results)
    print_lock = threading.Lock()
    
    import queue
    ebay_queue = queue.Queue()
    for item in unique_results:
        item["_retries"] = 0
        ebay_queue.put(item)

    def process_ebay_item(item):
        nonlocal comp_counter
        
        query          = item["ai"].get("ebay_search_query") or item["ai"].get("item_name", "")
        item_name      = item["ai"].get("item_name", "")
        ai_val_low     = float(item["ai"].get("ai_value_low", 0) or 0)
        fallback_query = item["ai"].get("ebay_fallback_query")
        ebay_condition = item["ai"].get("ebay_condition")
        inclusion_keywords = item["ai"].get("ebay_inclusion_keywords", [])
        exclusion_keywords = item["ai"].get("ebay_exclusion_keywords", [])

        comps_res = scrape_ebay_comps(
            None,
            query,
            ai_val_low,
            item_name,
            fallback_query=fallback_query,
            ebay_condition=ebay_condition,
            inclusion_keywords=inclusion_keywords,
            exclusion_keywords=exclusion_keywords,
        )
        
        # Requeue once on 0 results so that collateral victims of a soft-block get retried
        if comps_res["count"] == 0 and item.get("_retries", 0) < 1:
            item["_retries"] += 1
            with print_lock:
                print(f"  [Queue] 0 results for '{query}'. Requeuing to retry after potential cooldown.")
            ebay_queue.put(item)
            return

        item["comps"] = comps_res

        with print_lock:
            comp_counter += 1
            print(
                f"[{comp_counter}/{total_unique}] {query}\n"
                f"  -> {comps_res['low']} / {comps_res['mean']} / "
                f"{comps_res['high']} ({comps_res['count']} sales)"
            )

    def ebay_worker():
        while True:
            try:
                item = ebay_queue.get_nowait()
            except queue.Empty:
                break
                
            try:
                process_ebay_item(item)
            except Exception as exc:
                print(f"  eBay worker exception: {exc}")
            finally:
                ebay_queue.task_done()

    threads = []
    for _ in range(EBAY_WORKERS):
        t = threading.Thread(target=ebay_worker)
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    # --- Step 4: Generate report
    import urllib.parse
    from datetime import datetime

    sale_id = "Unknown"
    if url:
        try:
            path = urllib.parse.urlparse(url.strip()).path
            segments = [s for s in path.split("/") if s.isdigit()]
            if segments:
                sale_id = segments[-1]
        except Exception:
            pass

    current_time = datetime.now().strftime("%Y-%m-%d_%H%M")
    
    # Check if a custom report directory is provided in the environment
    from core.config import REPORT_OUTPUT_DIR, EMAIL_REPORTS
    
    if REPORT_OUTPUT_DIR:
        out_dir = Path(REPORT_OUTPUT_DIR)
    else:
        out_dir = Path(OUTPUT_FOLDER)
        
    # Ensure output folder exists
    out_dir.mkdir(parents=True, exist_ok=True)

    final_output_path = str(out_dir / f"EstateReport_{sale_id}_{current_time}.html")

    generate_report(unique_results, final_output_path, skipped_items=skipped_results)
    print(f"\nReport successfully saved to {final_output_path}")

    # --- Step 5: Generate Duplicates Excel report (if enabled)
    if GENERATE_DUPLICATES_REPORT:
        try:
            xlsx_path = str(out_dir / f"Duplicate_and_Probable_Duplicates_EstateReport_{sale_id}.xlsx")
            _generate_duplicates_xlsx(
                unique_results, merge_log, similar_flags, xlsx_path,
            )
            print(f"Duplicates report saved to {xlsx_path}")
        except Exception as exc:
            print(f"Warning: Failed to generate duplicates report: {exc}")

    if EMAIL_REPORTS:
        from core.email_sender import send_report_email
        send_report_email(final_output_path, url=url, items=unique_results)
    
    # Close the shared curl_cffi session to allow clean exit of the Python process
    close_ebay_session()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="AISaleAnalyst")
    parser.add_argument("--max-images", type=int, default=None, help="Temporarily override MAX_IMAGES from .env")
    parser.add_argument("--dev", action="store_true", help="Enable development testing mode (cheaper AI model, capped items, no emails)")
    args = parser.parse_args()
    main(max_images_override=args.max_images)